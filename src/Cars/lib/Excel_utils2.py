'''
Created on May 6, 2020
Excel utilities to open, read, and write to excel files in .xlsx format

'''
import openpyxl
from openpyxl.styles import Font

class Excel_utils2:

    def __init__(self, filename, sheet, in_out):
        self.filename = filename
        self.sheet = sheet
        self.in_out = in_out
        
        if (self.in_out == 'in'):
            self.wkbk = openpyxl.load_workbook(self.filename)
            self.sht = self.wkbk[self.sheet]
        elif (self.in_out == 'out'):
            self.wkbk = openpyxl.Workbook()
            self.sht = self.wkbk.active
            self.sht.title = self.sheet
            
    def set_cell(self, row, column, cell_value, cell_font, cell_bold, cell_size):
            self.sht.cell(row, column).value = cell_value 
            wc = self.sht.cell(row,column)
            wc.font = Font(name = cell_font, b=cell_bold, size = cell_size)
            
    def save_file(self, filename_out):
        self.filename_out = filename_out
        self.wkbk.save(self.filename_out)