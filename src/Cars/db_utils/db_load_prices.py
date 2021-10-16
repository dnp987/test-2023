'''
Created on Aug 21, 2020

@author: DNP Enterprises Inc.
Loads the carsdb.dealers database from spreadsheets
'''
import mysql.connector
import openpyxl

if __name__ == '__main__':
    
    mydb = mysql.connector.connect( # setup the database connection
    host = "localhost", 
    user = "dnp",
    password = "Npaapita5",
    database = "carsdb"
    )
  
    car_data_file = 'C:/Users/Home/Desktop/Cars/CarData.xlsx'
    dealer_wkbk = openpyxl.load_workbook(car_data_file)

    for index, dealer_sheet in enumerate(dealer_wkbk.sheetnames): #get each dealer sheet
        print (dealer_sheet)
        dealer_wksh = dealer_wkbk[dealer_sheet]
        for index, row in enumerate(dealer_wksh.rows): # for each sheet, scan the rows for the dealer price sheet file name
            if index == 0:
                continue # skip the first row
            dealer_price_file = row[6].value
            print (dealer_price_file)
            print ("------------------------------------------------------------------------------------------------------------------------")
            price_wkbk = openpyxl.load_workbook(dealer_price_file)
            record_count = 0
            price_wksht = price_wkbk.active
            for row in price_wksht.rows:
                dealer_id = row[0].value
                year = row[1].value
                make = row[2].value
                model = row[3].value
                price = row[4].value
                stock_num =row[5].value
                link = row[6].value
                print ("ID: ", dealer_id, "Year: ", year, "Make: ", make, "Model: ", model, "Price: ", price, "Stock #: ", stock_num, "URL: ", link)
                mycursor = mydb.cursor() 
                sql = 'INSERT into prices (dealer_id, year, make, model, price, stock_num, URL) VALUES (%s, %s, %s, %s, %s, %s, %s)'
                val = (dealer_id, year, make, model, price, stock_num, link)
                mycursor.execute(sql, val)
                mydb.commit()
                record_count +=1
            print ("Record count: ", record_count)
